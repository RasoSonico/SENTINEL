from datetime import date
from io import BytesIO
from typing import Literal

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from obra.models import Construction
from .data_service import PhysicalAdvanceRow, PhysicalAdvanceTableData


# ── Colour palette ──────────────────────────────────────────────────────────
_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
_SUBHDR_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
_ROW_ODD_FILL  = PatternFill("solid", fgColor="DCE6F1")   # light blue
_ROW_EVEN_FILL = PatternFill("solid", fgColor="FFFFFF")   # white
_TOTAL_FILL    = PatternFill("solid", fgColor="BDD7EE")   # soft blue

_WHITE_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_NORMAL_FONT = Font(name="Calibri", size=10)
_TOTAL_FONT  = Font(name="Calibri", bold=True, size=10)

_THIN_SIDE   = Side(style="thin", color="B0B0B0")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

# (col_index, header_label, width)
_COLUMNS_ALL = [
    (1,  "Partida",                        18),
    (2,  "Concepto",                       40),
    (3,  "Cantidad\ncontratada",           14),
    (4,  "Unidad",                          9),
    (5,  "Vol. ejecutado\ndel periodo",    16),
    (6,  "Avance del\nperiodo (%)",        14),
    (7,  "Importe\ndel periodo",           18),
    (8,  "Vol. total",                     16),
    (9,  "Importe\ntotal",                 18),
    (10, "% Avance\nacumulado",            14),
]

_COLUMNS_PERIOD = [
    (1, "Partida",                        18),
    (2, "Concepto",                       40),
    (3, "Cantidad\ncontratada",           14),
    (4, "Unidad",                          9),
    (5, "Vol. ejecutado\ndel periodo",    16),
    (6, "Avance del\nperiodo (%)",        14),
    (7, "Importe\ndel periodo",           18),
]


def _get_columns(scope: str):
    return _COLUMNS_ALL if scope == 'all' else _COLUMNS_PERIOD


def _apply_border_and_fill(cell, fill):
    cell.fill = fill
    cell.border = _THIN_BORDER


def _write_title(ws, construction: Construction, date_from: date, date_to: date, num_cols: int):
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"Reporte de Avance Físico — {construction.name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = _HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    period_cell = ws["A2"]

    if date_from == date_to:
        period_label = f"Avance del {date_from.strftime('%d/%m/%Y')}"
    else:
        period_label = f"Avance del {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}"

    period_cell.value = period_label
    period_cell.font = Font(name="Calibri", size=11, color="FFFFFF")
    period_cell.fill = _SUBHDR_FILL
    period_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6  # spacer


def _write_headers(ws, columns: list):
    ws.row_dimensions[4].height = 40
    for col_idx, header, _ in columns:
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = _WHITE_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER


def _write_data_rows(ws, rows: list[PhysicalAdvanceRow], scope: str) -> int:
    for row_idx, row in enumerate(rows, start=5):
        fill = _ROW_ODD_FILL if row_idx % 2 else _ROW_EVEN_FILL

        if scope == 'all':
            values = [
                row.work_item_name,
                row.concept_description,
                row.contracted_qty,
                row.unit,
                row.executed_in_range,
                row.period_pct / 100,
                row.importe_periodo,
                row.total_executed,
                row.importe_total,
                row.progress_pct / 100,
            ]
            number_formats = {
                3: '#,##0.00',
                5: '#,##0.00',
                6: '0.00%',
                7: '"$"#,##0.00',
                8: '#,##0.00',
                9: '"$"#,##0.00',
                10: '0.00%',
            }
        else:  # period
            values = [
                row.work_item_name,
                row.concept_description,
                row.contracted_qty,
                row.unit,
                row.executed_in_range,
                row.period_pct / 100,
                row.importe_periodo,
            ]
            number_formats = {
                3: '#,##0.00',
                5: '#,##0.00',
                6: '0.00%',
                7: '"$"#,##0.00',
            }

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _NORMAL_FONT
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_idx, fmt in number_formats.items():
            ws.cell(row=row_idx, column=col_idx).number_format = fmt

        ws.row_dimensions[row_idx].height = 18

    return 4 + len(rows)


def _write_totals(
    ws,
    rows: list[PhysicalAdvanceRow],
    last_data_row: int,
    scope: str,
    num_cols: int,
    catalog_total_value: float,
):
    total_row = last_data_row + 1
    ws.row_dimensions[total_row].height = 20

    # Style all cells in the row
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=total_row, column=col_idx)
        _apply_border_and_fill(cell, _TOTAL_FILL)
        cell.font = _TOTAL_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # Label spanning cols 1–2
    ws.cell(row=total_row, column=1, value="TOTALES").alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.merge_cells(
        start_row=total_row, start_column=1,
        end_row=total_row,   end_column=2,
    )

    imp_periodo = sum(r.importe_periodo for r in rows)

    # Avance del periodo (%) = Σ importe_periodo / Σ (qty × unit_price) all concepts
    period_pct = (imp_periodo / catalog_total_value) if catalog_total_value > 0 else 0.0
    pct_p_cell = ws.cell(row=total_row, column=6, value=period_pct)
    pct_p_cell.number_format = '0.00%'

    # Importe del periodo — col 7 in both scopes
    imp_p_cell = ws.cell(row=total_row, column=7, value=imp_periodo)
    imp_p_cell.number_format = '"$"#,##0.00'

    if scope == 'all':
        imp_total = sum(r.importe_total for r in rows)

        # Importe total — col 9
        imp_t_cell = ws.cell(row=total_row, column=9, value=imp_total)
        imp_t_cell.number_format = '"$"#,##0.00'

        # % Avance acumulado = Σ importe_total / Σ (qty × unit_price) all concepts
        progress_pct = (imp_total / catalog_total_value) if catalog_total_value > 0 else 0.0
        pct_a_cell = ws.cell(row=total_row, column=10, value=progress_pct)
        pct_a_cell.number_format = '0.00%'


def _set_column_widths(ws, columns: list):
    for col_idx, _, width in columns:
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def generate_physical_advance_excel(
    construction: Construction,
    table_data: PhysicalAdvanceTableData,
    date_from: date,
    date_to: date,
    scope: Literal['all', 'period'] = 'all',
) -> BytesIO:
    columns = _get_columns(scope)
    num_cols = len(columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avance Físico"
    ws.freeze_panes = "A5"

    _write_title(ws, construction, date_from, date_to, num_cols)
    _write_headers(ws, columns)
    last_data_row = _write_data_rows(ws, table_data.rows, scope)

    if table_data.rows:
        _write_totals(
            ws,
            table_data.rows,
            last_data_row,
            scope,
            num_cols,
            table_data.catalog_total_value,
        )

    _set_column_widths(ws, columns)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
